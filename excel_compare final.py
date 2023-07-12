import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import tkinter as tk
from tkinter import filedialog

def compare_excel_files(master_file_path, compare_file_path, output_file_path):
    
    master_df = pd.read_excel(master_file_path, header=None)
    compare_df = pd.read_excel(compare_file_path, header=None)
    
    diff_book = Workbook()
    diff_sheet = diff_book.active
    
    for row in range(compare_df.shape[0]):
        for col in range(compare_df.shape[1]):
            compare_cell = compare_df.iloc[row, col]
            master_cell = master_df.iloc[row, col]
            
    
            if compare_cell != master_cell:
    
                font = Font(color="FF0000", bold=True)  # Red font color
                fill = PatternFill(fill_type="solid", fgColor="FFFF00")  # Yellow fill color
                
    
                diff_sheet.cell(row=row+1, column=col+1, value=compare_cell)
                diff_sheet.cell(row=row+1, column=col+1).font = font
                diff_sheet.cell(row=row+1, column=col+1).fill = fill
            else:

                diff_sheet.cell(row=row+1, column=col+1, value=compare_cell)
    
    diff_book.save(output_file_path)

root = tk.Tk()
root.withdraw()  # Hide the main window

# Prompt the user to select the master Excel file
master_file_path = filedialog.askopenfilename(title="Select the master Excel file")

# Prompt the user to select the Excel file to compare against
compare_file_path = filedialog.askopenfilename(title="Select the Excel file to compare against")

# Prompt the user to choose the output file path
output_file_path = filedialog.asksaveasfilename(title="Save the output Excel file")

compare_excel_files(master_file_path, compare_file_path, output_file_path)
